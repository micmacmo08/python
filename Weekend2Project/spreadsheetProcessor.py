from openpyxl import load_workbook
from re import split

goodOrBad = lambda x, min: ("Good" if x >= min else "Bad")

months = {
     'january': (1, 'january', 'jan'),
     'february': (2,'february', 'feb'), 
     'march': (3,'march', 'mar'), 
     'april': (4,'april', 'apr'), 
     'may': (5,'may'), 
     'june':(6, 'june', 'jun'), 
     'july':(7, 'july', 'jul'), 
     'august':(8, 'august', 'aug'), 
     'september':(9, 'september', 'sep'), 
     'october':(10, 'october', 'oct'), 
     'november':(11, 'november', 'nov'), 
     'december':(12, 'december', 'dec')}
promoters = 0
passives = 0
dectractors = 0
promoterMin = 200
passiveMin = 100
dectractorMin = 100

workingDir = 'F:\\Users\\Mike\\OneDrive\\Documents\\GitHub\\Python\\Weekend2Project\\'


def processSheet(file, logger):
    wb = load_workbook(workingDir + file)
    ws = wb["Summary Rolling MoM"]

    month, year = split('expedia_report_monthly_|_|.xlsx', file)[1:-1]
    data = ()

    headers = [cell.value for cell in ws[1] if cell.value is not None]
    try:
        for row in ws.iter_rows(min_row = 2, values_only=True): #iterate through the rows, starting with the second row, to find the row with the correct month and save that data
            r = []
            for cell in row:
                if cell is None: #ignore empty cells
                    continue
                else:
                    r.append(cell)
            if r and r[0].month in months[month.lower()]: 
                data = (tuple(r))
                break #once we find the correct row, we can stop searching

    except:
        logger.info("Month not found in 'Summary Rolling MoM'")
        return 'Error'

    logger.info("------------------------------")
    logger.info("MONTH: {}".format(month.upper()))
    for i in range(len(headers)):
        #cut off white space at beginning or end of headers
        if headers[i][0] == ' ':
            headers[i] = headers[i][1:]
        if headers[i][-1] == ' ':
            headers[i] = headers[i][:-1]
        if type(data[i+1]) is int:
            logger.info("{}: {:,}".format(headers[i], data[i+1])) #data is offset by one due to there not being a header for the month
        else:
            logger.info("{}: {:.2%}".format(headers[i], data[i+1])) #data is offset by one due to there not being a header for the month
    
    ws = wb["VOC Rolling MoM"]
    index = 0
    found = False
    for cell in ws[1]:
        if cell.value is not None and type(cell.value) != str and cell.value.month in months[month]:
            found = True
            break
        elif cell.value is not None and cell.value in months[month]:
            found = True
            break
        index += 1
    if not found:
        logger.info('Month not found in VOC Rolling MoM')
        return 'Error'

    promoFound = False
    passFound = False
    dectractFound = False
    for row in ws.iter_rows(min_row = 2, values_only = True):
        if row[0] is not None and 'promoters' in row[0].lower():
            promoters = row[index]
            promoFound = True
        elif row[0] is not None and 'passives' in row[0].lower():
            passives = row[index]
            passFound = True
        elif row[0] is not None and 'dectractors' in row[0].lower():
            dectractors = row[index]
            dectractFound = True
    
    if not promoFound:
        logger.info('Promotors not found in VOC Rolling MoM')
        return 'Error'
    if not passFound:
        logger.info('Passives not found in VOC Rolling MoM')
        return 'Error'
    if not dectractFound:
        logger.info('Dectractors not found in VOC Rolling MoM')
        return 'Error'
    
    logger.info("Promoters: {} ({})".format(promoters, goodOrBad(promoters,promoterMin)))
    logger.info("Passives: {} ({})".format(passives, goodOrBad(passives,passiveMin)))
    logger.info("Dectractors: {} ({})".format(dectractors, goodOrBad(dectractorMin,promoterMin)))
    logger.info("------------------------------")

    logger.info('{} successfully processed'.format(file))
    return 'Archive'


