import logging
import re
import openpyxl
import sys

months = {'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june':6, 'july':7, 'august':8, 'september':9, 'october':10, 'november':11, 'december':12}

logging.basicConfig(filename='project.log', level=logging.INFO)

#get user input and loop until successful
while True:
    try:
        uInput = input("Please enter the month for which you want the data and the filename from which you'd like the data using the format [month]([filename]) ex January(expedia_report_monthly_january_2018.xlsx):\n")
        uInput = re.split('\(|\)', uInput)
        month = uInput[0].lower()
        filename = uInput[1]

        #Open file and read contents
        wb = openpyxl.load_workbook(filename)
        break
    except:
        print("\nIncorrect input, try again\n")

ws = wb["Summary Rolling MoM"]
headers = [cell.value for cell in ws[1] if cell.value is not None] #grab the headers from the first row
data = ()


try:
    for row in ws.iter_rows(min_row = 2, values_only=True): #iterate through the rows, starting with the second row, to find the row with the correct month and save that data
        r = []
        for cell in row:
            if cell is None: #ignore empty cells
                continue
            else:
                r.append(cell)
        if r and r[0].month == months[month]: 
            data = (tuple(r))
            break #once we find the correct row, we can stop searching
except:
    print("\nMonth not found in sheet. Check data in sheet and run program again.\n")
    sys.exit(1)

#log the data
logging.info("MONTH: {}".format(month.upper()))
for i in range(len(headers)):
    #cut off white space at beginning or end of headers
    if headers[i][0] == ' ':
        headers[i] = headers[i][1:]
    if headers[i][-1] == ' ':
        headers[i] = headers[i][:-1]
    if type(data[i+1]) is int:
        logging.info("{}: {:,}".format(headers[i], data[i+1])) #data is offset by one due to there not being a header for the month
    else:
        logging.info("{}: {:.2%}".format(headers[i], data[i+1])) #data is offset by one due to there not being a header for the month


